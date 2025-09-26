#PASO 1 INTALALR.- pipenv install openpyxl
import openpyxl

wb = openpyxl.load_workbook("13-paquetes-populares/yucra.xlsx")

#print(wb.sheetnames) #imprime las hojas del excel
#print(wb["Ejercicio 1"]) #imprime la hoja que l digas del excel

hoja = wb.active #esto sirve para seleccionar la hoja activa del excel
# print(hoja)

wb.create_sheet("Ejercicio 5") #esto sirve para crear una hoja nueva en el excel
hoja5 = wb["Ejercicio 5"] #esto sirve para seleccionar la hoja que acabde de crear
#hoja5.title = "Ejercicio 55" #esto sirve para cambiar el nombre de la hoja "

# print(hoja)
# print(
#     hoja.title,
#     hoja.max_row,
#     hoja.max_column
# )

celda = hoja["A1"] #esto sirve para seleccionar una celda del excel
celda.value = "Hola Mundo mundano" #esto sirve para asignar un valor a la celda seleccionada
#print(celda.value) #esto sirve para imprimir el valor de la celda seleccionada

# celda2 = hoja.cell(row=5, column=2) #esto sirve para seleccionar una celda del excel por fila y columna
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
    # ) #esto sirve para imprimir el valor de la celda seleccionada

# for fila in range(1, hoja.max_row +1):
#     for columna in range(1, hoja.max_column +1):
#         celda = hoja.cell(row=fila, column=columna)
#         print(celda.coordinate, celda.value)
# 
columna = hoja["A"]
fila = hoja["1"]
#print(columna)
#print(fila)

hoja.append([1, 2, 3])
print(hoja.rows)
hoja.delete_rows(6, 1) # los parametros indican desde que fila y cuantas filas borrar
wb.save("13-paquetes-populares/yucra_mod.xlsx") #esto sirve para guardar los cambios en el excel




# # Imprimir como matriz visual
# for fila in range(1, hoja.max_row + 1):
#     fila_datos = []
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)
#         # Convertir el valor a string y alinear
#         valor = str(celda.value) if celda.value is not None else ""
#         fila_datos.append(f"{valor:15}")  # Ancho fijo de 15 caracteres
    
#     # Imprimir toda la fila en una sola línea
#     print(" | ".join(fila_datos))